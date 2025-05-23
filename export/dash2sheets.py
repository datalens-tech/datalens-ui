# "python3 dash2sheets.py OUTPUT_NAME=output.xlsx SEP=; ./temp/file1.csv ./temp/file2.csv"
#
# Python скрипт для объединения нескольких CSV файлов в один XLSX
# 
# Параметры:
# - OUTPUT_NAME: string - выходной файл
# - SEP: string - разделитель CSV 
# - далее следует список CSV файлов

import pandas as pd
import sys
import os
import openpyxl
from copy import copy

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

def read_args():
    """
    Чтение аргументов из командной строки

    Результат:
    ----------
    Словарь с параметрами
    """

    params = {}

    new_argv = sys.argv[1:5]

    for i in range(0, len(new_argv)):
        arg = new_argv[i].split('=')
        
        if len(arg) > 1:
            params[arg[0].replace('\"', '')] = arg[1].replace('\"', '')

    return params

params = read_args()

meta = open(params['META_NAME'], "r").readlines()
headersDir = params['HEADERS_PATH']

filenames = sys.argv[5:]

wb = openpyxl.Workbook()   
wb.remove(wb.active)

for id, csvfilename in enumerate(filenames):
    sheetName = meta[id].replace('\n', '')[:31]
    sheet = wb.create_sheet(sheetName)
    
    headerPath = headersDir + '/' + sheetName + '.xlsx'

    df = pd.read_csv(csvfilename, sep=params["SEP"])

    if os.path.exists(headerPath):
        header = openpyxl.load_workbook(headerPath)
        headerSheet = header[header.sheetnames.pop()]

        copy_sheet(headerSheet, sheet)
    else:
        sheet.append(df.columns.to_list())

    for row in df.values:
        sheet.append(row.tolist())

wb.save(params["OUTPUT_NAME"])